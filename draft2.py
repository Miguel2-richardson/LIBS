# -*- coding: utf-8 -*-
"""
Created on Mon Mar 15 10:47:28 2021

@author: Miguel
"""
#Trying a different approach
from spectraasc import dataread, spectraread, getpeaks, featuresread, featurecompare
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as px
import tensorflow as tf
from tensorflow import keras
'''
def insert_pos(position, list1, list2):
    return list1[:position] + list2 + list1[position:]

spec = 2
trig = 0.03
file = ["Nist1.xlsx", "Nist2.xlsx", "Nist26.xlsx",] 
# f1 = ["Pure Aluminum_1.asc"]
# l1,d1 = spectraread(f1[0])
Data = []
Wavelen = []
hld = []
#Reading in data into 1 list, same for wavelen
for i in file:
    wavelen, data = dataread(i)
    Data.append(data)
    Wavelen.append(wavelen)
    hld.append(len(wavelen))

# This is the correct way to interact with the data and be able to do thing with
#the peak information after.

w = np.array(Wavelen[spec])
current = np.array(Data[spec])
peaks, height, left, right = getpeaks(current, trig)

#End of method
plt.plot(w[peaks],current[peaks], "x")
plt.plot(Wavelen[spec],Data[spec])
print(len(peaks))

for i in range(0, len(file)):
    print(i)



Q = [[0,1,2],[3,4,5],[6,7,8]]
Q =np.array(Q)
print(Q)

print(np.arange(0,1,0.001))
'''

#Creating feature array
#Open excel file with features
#compare positions of features in excel file with positions of features found
#store intensities in wavelength order into a numpy array
#return this numpy array
'''
featuresFile = ["Features.xlsx"]

def featuresread(file):
    #loading excelsheet
    wb = px.load_workbook(file)
    ws_name = wb.sheetnames
    p1_name = ws_name[0]
    #end sheet load
    
    p1 = wb[p1_name]
    
        
    opensheet = wb.active
    lastrow = opensheet.max_row
    lastrow = ["A" + str(lastrow) ]
    
    features = []

    
    for i in p1['A2':lastrow[0]]:
        for k in i:
            features.append(float(k.internal_value))

    wb.close()



       
    return(features)
#THis accomplishes the goal of reading in the features from the excel file.
#I want to functionalize it, so calling it returns the features list without
#needing to provide the filename.
#def features():
    #file = ["Features.xlsx]

F1 = featuresread(featuresFile[0])
'''

'''
def featuresread():
    #loading excelsheet
    file = ["Features.xlsx"]
    wb = px.load_workbook(file[0])
    ws_name = wb.sheetnames
    p1_name = ws_name[0]
    #end sheet load
    
    p1 = wb[p1_name]
    
        
    opensheet = wb.active
    lastrow = opensheet.max_row
    lastrow = ["A" + str(lastrow) ]
    
    features = []

    
    for i in p1['A2':lastrow[0]]:
        for k in i:
            features.append(float(k.internal_value))

    wb.close()
    features = np.array(features)   
    return(features)

F1 = featuresread()
'''
#Reading in data and using find peaks
file = ["Nist1.xlsx", "Nist2.xlsx", "Nist3.xlsx","Nist4.xlsx","Nist5.xlsx",
        "Nist6.xlsx","Nist7.xlsx","Nist8.xlsx","Nist9.xlsx","Nist10.xlsx",
        "Nist11.xlsx","Nist12.xlsx","Nist13.xlsx","Nist14.xlsx","Nist15.xlsx",
        "Nist16.xlsx","Nist17.xlsx","Nist18.xlsx","Nist19.xlsx","Nist20.xlsx"
        ,"Nist21.xlsx","Nist22.xlsx","Nist23.xlsx","Nist24.xlsx","Nist25.xlsx",
        "Nist26.xlsx","Nist27.xlsx","Nist28.xlsx","Nist30.xlsx","Nist31.xlsx",
        "Nist32.xlsx","Nist33.xlsx","Nist34.xlsx","Nist35.xlsx","Nist36.xlsx",
        "Nist37.xlsx","Nist38.xlsx","Nist39.xlsx","Nist40.xlsx","Nist41.xlsx",
        "Nist42.xlsx","Nist43.xlsx","Nist44.xlsx","Nist45.xlsx","Nist46.xlsx",
        "Nist47.xlsx","Nist48.xlsx","Nist49.xlsx","Nist50.xlsx","Nist51.xlsx",
        "Nist52.xlsx","Nist53.xlsx","Nist54.xlsx","Nist55.xlsx","Nist56.xlsx",
        "Nist57.xlsx","Nist58.xlsx","Nist59.xlsx","Nist60.xlsx","Nist61.xlsx",
        "Nist62.xlsx","Nist63.xlsx","Nist64.xlsx","Nist65.xlsx","Nist66.xlsx",
        "Nist67.xlsx","Nist68.xlsx","Nist69.xlsx","Nist70.xlsx","Nist71.xlsx",
        "Nist72.xlsx","Nist73.xlsx","Nist74.xlsx","Nist75.xlsx","Nist76.xlsx",
        "Nist77.xlsx","Nist78.xlsx","Nist79.xlsx","Nist80.xlsx","Nist81.xlsx",
        "Nist82.xlsx","Nist83.xlsx"]

 
Data = []
Wavelen = []
Label = []
trig = 0.03

peaks = []
#Reading in data into 1 list, same for wavelen
for i in file:
    wavelen, data, label = dataread(i)
    Data.append(data)
    Wavelen.append(wavelen)
    Label.append(label)


# print(len(Wavelen))
# D1 = np.array(Data[0])
# W1 = np.array(Wavelen[0])
'''
peaks, height, left, right = getpeaks(Data[0],trig )

W1 = Wavelen[0]
W1 = [W1[i] for i in peaks]
print(W1)
'''
# plt.plot(W1,D1)
# plt.plot(W1[peaks],D1[peaks], "x")




'''
def featurecompare(wavelen, data, trig):
    features = featuresread()
    foundInten = []
    foundWave = []

    nnfeatures = []
    for i in range(len(data)):
        peaks, h, l, r = getpeaks(data[i], trig)
        d,w = data[i], wavelen[i] #data and wavelen of current spectra stored
        d = [d[l] for l in peaks] #intensity of the found features 
        w = [w[l] for l in peaks] #wavelen of the found features
        
        foundInten.append(d)
        foundWave.append(w)
    #2 list of lists. One with the intensity of the features for each spectra.
    #and the corresponding wavelengths
    for i in range(len(foundWave)):
        nnspectra = [0]*len(features)
        for f in features:
            for w in foundWave[i]:
                if abs(f-w) <0.05:
                    holder = foundWave[i].index(w)
                    featureholder = features.index(f)
                    # nnspectra.insert(featureholder,foundInten[i][holder])
                    nnspectra[featureholder] = foundInten[i][holder]
                    print(f, " ",w) #check what peaks were found from the comparison
                    # print(len(features))
                # if (abs(f-w) > 0.05 and logic != "True"):
                    # nnspectra[features.index(f)] = 0
                    # nnspectra.insert(features.index(f), 0)
    
    
        nnfeatures.append(nnspectra)
    nnfeatures = np.array(nnfeatures)
    return(nnfeatures)
'''

stuff = featurecompare(Wavelen, Data, trig)
shape = len(stuff[0])
Label = [i/100 for i in Label]
output = 23
second = 2
third = 3
model = keras.Sequential([
    keras.layers.InputLayer(input_shape = (shape,)),
    keras.layers.Dense(second*shape, activation = "relu"),
    keras.layers.Dense(third*shape, activation = "relu"),
    keras.layers.Dense(output, activation = "linear")
    ])
model.compile(optimizer = 'adam',
              loss = tf.keras.losses.MeanSquaredError(),
              metrics = tf.keras.metrics.RootMeanSquaredError()
              )
model.fit(stuff, Label, epochs = 400)

model.save("test.h5")
