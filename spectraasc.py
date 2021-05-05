# -*- coding: utf-8 -*-
"""
Created on Wed Feb 10 14:44:45 2021

@author: Miguel
"""

# spectraasc.py>



from scipy.signal import find_peaks, peak_widths
#import matplotlib.pyplot as plt
import numpy as np
import openpyxl as px

"Takes an asc file and returns the wavelength and intensity of the spectra"
def spectraread(file):
    data = []
    with open(file, 'r') as f:
        for line in f:
            line = line.strip() #Remove \n from the line being read
            columns = line.split(",") #Removes , from line being read

            data.append(columns) #Appends remainder to the data list
    f.close()
    
    wavelen = [float(data[n][0]) for n in range(len(data))] #Changes wavelength to floats
    Intensity = [float(data[n][1]) for n in range(len(data))] #Changes intensity to floats
    
    datamin = min(Intensity)
    delta = max(Intensity) - datamin
    Intensity = [(i - datamin)/delta for i in Intensity] # data scaled
  
    return(wavelen, Intensity)


"Finds the peaks in the data and finds left and right interpolated points"
def getpeaks(Intensity, trigger):
    
    peaks, _ = find_peaks(Intensity, trigger)
    npIntens = np.array(Intensity)
    
    results = peak_widths(npIntens, peaks, rel_height=0.3)

    #results[1] heights of the contour lines where widths were evaluated. [2] left, [3 right]
    return(peaks,results[1],  results[2], results[3])


"Takes an excel sheet and returns wavelength and intensity"
def dataread(file):
    "loading in excel sheets"
    wb = px.load_workbook(file)
    ws_name = wb.sheetnames
    p1_name = ws_name[0] #Page 1 of excel sheet
    #p2_name = ws_name[1] #Page 2 of excel sheet
    "End excel sheets data load"
    
    p1 = wb[p1_name]
    
    opensheet = wb.active
    lastrow = opensheet.max_row
    lastrow = ["A" + str(lastrow), "B" + str(lastrow) ]
    
    wavelen = []
    rawdata = []
    
    for i in p1['A1':lastrow[0]]:
        for k in i:
            wavelen.append(float(k.internal_value))
    for i in p1['B1':lastrow[1]]:
        for k in i:
            rawdata.append(float(k.internal_value))
    label = p1['D2'].value
    wb.close()
    #label for % of the element being tracked
    
    datamin = min(rawdata)
    delta = max(rawdata) - datamin
    scaledata = [(i-datamin)/delta for i in rawdata]       
    return(wavelen, scaledata, label)

def widthcorrec(x_values, leftpoint, rightpoint):
    
    xmin = [x_values[int(i)] for i in leftpoint] # associating width position to wavelength
    xmax = [x_values[int(i)] for i in rightpoint] #Same as above for right endpoint of line
    # for i in range(0,len())
    width = [xmax[i] - xmin[i] for i in range(0,len(xmax))]
    return(width)

def featurecompare(wavelen, data, trig):
    features = featuresread()
    foundInten = []
    foundWave = []

    nnfeatures = []
    for i in range(len(data)):
        peaks, h, l, r = getpeaks(data[i], trig)
        d,w = data[i], wavelen[i] #data and wavelen of current spectra stored
        d = [d[l] for l in peaks] #intensity of the found features 
        w = [w[l] for l in peaks] #wavelen of the found features
        
        foundInten.append(d)
        foundWave.append(w)
    #2 list of lists. One with the intensity of the features for each spectra.
    #and the corresponding wavelengths
    for i in range(len(foundWave)):
        nnspectra = [0]*len(features)
        for f in features:
            for w in foundWave[i]:
                if abs(f-w) <0.05:
                    holder = foundWave[i].index(w)
                    featureholder = features.index(f)
                    # nnspectra.insert(featureholder,foundInten[i][holder])
                    nnspectra[featureholder] = foundInten[i][holder]
                    # print(f, " ",w) #check what peaks were found from the comparison
                    # print(len(features))
                # if (abs(f-w) > 0.05 and logic != "True"):
                    # nnspectra[features.index(f)] = 0
                    # nnspectra.insert(features.index(f), 0)
    
    
        nnfeatures.append(nnspectra)
    # nnfeatures = np.array(nnfeatures)
    return(nnfeatures)



def featuresread():
    #loading excelsheet
    file = ["Features2.xlsx"]
    wb = px.load_workbook(file[0])
    ws_name = wb.sheetnames
    p1_name = ws_name[0]
    #end sheet load
    
    p1 = wb[p1_name]
    
        
    opensheet = wb.active
    lastrow = opensheet.max_row
    lastrow = ["A" + str(lastrow) ]
    
    features = []

    
    for i in p1['A2':lastrow[0]]:
        for k in i:
            features.append(float(k.internal_value))

    wb.close()
    #features = np.array(features)   
    return(features)




def nistformat(Wavelen, Data,file, hld):
    #accepts data and the number of files. Returns a numpy array of data
    Max = [max(i) for i in Wavelen]
    Max = max(Max) 
    step = max(hld)
    newWavelen = []
    pos = [i for i in range(max(hld))] #Position tracker
    for i in range(file):
        counter = []
        m = np.diff(Wavelen[i])
        diff = np.diff(m)
        for t in diff:
            if t <0.1:
                t = 0
            counter.append(t)
        counter = [int(a/m[0]) for a in counter]
    
        for b in counter:
            if b != 0:
                it = [0]*b
                Data[i] = Data[i][:counter.index(b)+2] + it + Data[i][counter.index(b)+2:]
                counter = counter[:counter.index(b)+2] + it +counter[counter.index(b)+2:]
        end = len(pos) - len(Data[i])
        end = [0]*end
        Data[i] = Data[i] + end
        # newWavelen[i] = np.linspace(Wavelen[i][0], Max, num = step )
        newWavelen.append(np.linspace(Wavelen[i][0], Max, num = step ))
    Data = np.array(Data)
    newWavelen = np.array(newWavelen)
    return(newWavelen, Data)